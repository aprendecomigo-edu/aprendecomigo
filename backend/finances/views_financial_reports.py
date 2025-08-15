"""
Financial Reports API Views for Analytics and Export functionality.

This module provides comprehensive financial reporting capabilities including:
- Student spending analytics
- School financial overview
- Teacher compensation reports
- Revenue trend analysis
- Data export functionality (CSV, Excel, PDF)
- Receipt generation and management
"""

import csv
from datetime import datetime, timedelta
from decimal import Decimal
import logging

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None


from django.contrib.auth import get_user_model
from django.db.models import Count, Sum
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import permissions, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.permissions import IsSchoolOwnerOrAdmin

from .models import (
    ClassSession,
    PurchaseTransaction,
    Receipt,
    StudentAccountBalance,
    TransactionPaymentStatus,
    TransactionType,
)
from .serializers import (
    ReceiptGenerationRequestSerializer,
    ReceiptSerializer,
)
from .services.receipt_service import ReceiptGenerationService

User = get_user_model()
logger = logging.getLogger(__name__)


class FinancialAnalyticsAPIView(APIView):
    """API views for financial analytics and reporting."""

    permission_classes = [permissions.IsAuthenticated, IsSchoolOwnerOrAdmin]

    def _get_user_schools(self):
        """Get schools that the user has access to."""
        from accounts.models import School, SchoolRole

        # Admin users can see all schools
        if self.request.user.is_staff or self.request.user.is_superuser:
            return School.objects.all()

        # Get schools where user is owner or admin
        user_roles = SchoolRole.objects.filter(user=self.request.user, role__in=["owner", "admin"])
        return School.objects.filter(id__in=user_roles.values_list("school_id", flat=True))


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def student_spending_analytics(request: Request) -> Response:
    """
    Get spending analytics for the authenticated student.

    Query parameters:
    - period: 'monthly', 'quarterly', 'yearly' (default: 'monthly')
    """
    try:
        # Only students can access their own analytics
        if not hasattr(request.user, "account_balance"):
            return Response({"error": "Only students can access spending analytics"}, status=status.HTTP_403_FORBIDDEN)

        period = request.query_params.get("period", "monthly")

        # Calculate date range based on period
        end_date = timezone.now()
        if period == "monthly":
            start_date = end_date - timedelta(days=30)
        elif period == "quarterly":
            start_date = end_date - timedelta(days=90)
        elif period == "yearly":
            start_date = end_date - timedelta(days=365)
        else:
            start_date = end_date - timedelta(days=30)

        # Get completed transactions for the student
        transactions = PurchaseTransaction.objects.filter(
            student=request.user, payment_status=TransactionPaymentStatus.COMPLETED, created_at__gte=start_date
        )

        # Calculate analytics
        total_spending = transactions.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

        transaction_count = transactions.count()

        # Spending by transaction type
        spending_by_type = {}
        for transaction_type in TransactionType.choices:
            type_total = transactions.filter(transaction_type=transaction_type[0]).aggregate(total=Sum("amount"))[
                "total"
            ] or Decimal("0.00")
            spending_by_type[str(transaction_type[1])] = float(type_total)

        # Monthly breakdown for the period
        monthly_breakdown = []
        current_date = start_date
        while current_date <= end_date:
            month_start = current_date.replace(day=1)
            next_month = (month_start + timedelta(days=32)).replace(day=1)

            month_spending = transactions.filter(created_at__gte=month_start, created_at__lt=next_month).aggregate(
                total=Sum("amount")
            )["total"] or Decimal("0.00")

            monthly_breakdown.append({"month": month_start.strftime("%Y-%m"), "spending": float(month_spending)})

            current_date = next_month
            if current_date > end_date:
                break

        return Response(
            {
                "total_spending": float(total_spending),
                "transaction_count": transaction_count,
                "spending_by_type": spending_by_type,
                "monthly_breakdown": monthly_breakdown,
                "period": period,
                "date_range": {"start": start_date.date().isoformat(), "end": end_date.date().isoformat()},
            }
        )

    except Exception as e:
        logger.error(f"Student spending analytics error: {e}")
        return Response(
            {"error": "Failed to generate spending analytics"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated, IsSchoolOwnerOrAdmin])
def school_financial_overview(request: Request) -> Response:
    """
    Get school-wide financial overview for admin users.
    Requires school admin permissions.
    """
    try:
        from accounts.models import School, SchoolRole

        # Get schools accessible to the user
        if request.user.is_staff or request.user.is_superuser:
            schools = School.objects.all()
        else:
            user_roles = SchoolRole.objects.filter(user=request.user, role__in=["owner", "admin"])
            schools = School.objects.filter(id__in=user_roles.values_list("school_id", flat=True))

        if not schools.exists():
            return Response({"error": "No schools accessible"}, status=status.HTTP_403_FORBIDDEN)

        # Get all completed transactions for accessible schools
        # Note: This assumes transactions are linked to schools via students
        school_students = User.objects.filter(school_memberships__school__in=schools).distinct()

        transactions = PurchaseTransaction.objects.filter(
            student__in=school_students, payment_status=TransactionPaymentStatus.COMPLETED
        )

        # Calculate overview metrics
        total_revenue = transactions.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

        active_students = (
            school_students.filter(purchase_transactions__payment_status=TransactionPaymentStatus.COMPLETED)
            .distinct()
            .count()
        )

        transaction_count = transactions.count()

        # Revenue by transaction type
        revenue_by_type = {}
        for transaction_type in TransactionType.choices:
            type_revenue = transactions.filter(transaction_type=transaction_type[0]).aggregate(total=Sum("amount"))[
                "total"
            ] or Decimal("0.00")
            revenue_by_type[str(transaction_type[1])] = float(type_revenue)

        # Top students by spending
        top_students = (
            transactions.values("student__name", "student__email")
            .annotate(total_spent=Sum("amount"))
            .order_by("-total_spent")[:10]
        )

        return Response(
            {
                "total_revenue": float(total_revenue),
                "active_students": active_students,
                "transaction_count": transaction_count,
                "revenue_by_type": revenue_by_type,
                "top_students": list(top_students),
                "schools_count": schools.count(),
            }
        )

    except Exception as e:
        logger.error(f"School financial overview error: {e}")
        return Response({"error": "Failed to generate school overview"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated, IsSchoolOwnerOrAdmin])
def teacher_compensation_report(request: Request) -> Response:
    """
    Get teacher compensation calculation and reporting.

    Query parameters:
    - teacher_id: ID of specific teacher (required)
    - period: 'monthly', 'quarterly' (default: 'monthly')
    """
    try:
        teacher_id = request.query_params.get("teacher_id")
        if not teacher_id:
            return Response({"error": "teacher_id parameter is required"}, status=status.HTTP_400_BAD_REQUEST)

        period = request.query_params.get("period", "monthly")

        # Calculate date range
        end_date = timezone.now()
        start_date = end_date - timedelta(days=30) if period == "monthly" else end_date - timedelta(days=90)

        # Get teacher's completed sessions
        sessions = ClassSession.objects.filter(teacher_id=teacher_id, status="completed", date__gte=start_date.date())

        sessions_completed = sessions.count()

        # Calculate total hours taught (assuming 1 hour per session for now)
        total_hours_taught = sessions_completed * 1.0

        # Base compensation calculation (this should be based on actual compensation rules)
        # For now, using a default rate
        hourly_rate = Decimal("25.00")  # This should come from TeacherCompensationRule
        total_compensation = total_hours_taught * float(hourly_rate)

        # Hourly breakdown by date
        hourly_breakdown = (
            sessions.values("date")
            .annotate(hours_taught=Count("id"), daily_compensation=Count("id") * float(hourly_rate))
            .order_by("date")
        )

        return Response(
            {
                "teacher_id": int(teacher_id),
                "total_hours_taught": total_hours_taught,
                "total_compensation": total_compensation,
                "sessions_completed": sessions_completed,
                "hourly_breakdown": list(hourly_breakdown),
                "period": period,
                "hourly_rate": float(hourly_rate),
            }
        )

    except Exception as e:
        logger.error(f"Teacher compensation report error: {e}")
        return Response(
            {"error": "Failed to generate compensation report"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated, IsSchoolOwnerOrAdmin])
def revenue_trends_analysis(request: Request) -> Response:
    """
    Get revenue trends over time analysis.

    Query parameters:
    - period: 'daily', 'weekly', 'monthly' (default: 'daily')
    - days: number of days to analyze (default: 30)
    """
    try:
        period = request.query_params.get("period", "daily")
        days = int(request.query_params.get("days", 30))

        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)

        from accounts.models import School, SchoolRole

        # Get accessible schools
        if request.user.is_staff or request.user.is_superuser:
            schools = School.objects.all()
        else:
            user_roles = SchoolRole.objects.filter(user=request.user, role__in=["owner", "admin"])
            schools = School.objects.filter(id__in=user_roles.values_list("school_id", flat=True))

        # Get students from accessible schools
        school_students = User.objects.filter(school_memberships__school__in=schools).distinct()

        # Get completed transactions
        transactions = PurchaseTransaction.objects.filter(
            student__in=school_students, payment_status=TransactionPaymentStatus.COMPLETED, created_at__gte=start_date
        )

        # Calculate daily revenue
        daily_revenue = []
        current_date = start_date.date()
        end_date_date = end_date.date()

        while current_date <= end_date_date:
            day_transactions = transactions.filter(created_at__date=current_date)
            day_revenue = day_transactions.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

            daily_revenue.append(
                {
                    "date": current_date.isoformat(),
                    "revenue": float(day_revenue),
                    "transaction_count": day_transactions.count(),
                }
            )

            current_date += timedelta(days=1)

        # Calculate growth rate
        if len(daily_revenue) >= 2:
            recent_avg = sum([day["revenue"] for day in daily_revenue[-7:]]) / 7
            earlier_avg = sum([day["revenue"] for day in daily_revenue[:7]]) / 7
            growth_rate = ((recent_avg - earlier_avg) / max(earlier_avg, 1)) * 100
        else:
            growth_rate = 0.0

        # Find peak days
        peak_days = sorted(daily_revenue, key=lambda x: x["revenue"], reverse=True)[:3]

        # Total period revenue
        total_period_revenue = sum([day["revenue"] for day in daily_revenue])

        return Response(
            {
                "daily_revenue": daily_revenue,
                "growth_rate": growth_rate,
                "peak_days": peak_days,
                "total_period_revenue": total_period_revenue,
                "period": period,
                "days_analyzed": days,
            }
        )

    except Exception as e:
        logger.error(f"Revenue trends analysis error: {e}")
        return Response({"error": "Failed to generate revenue trends"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class FinancialExportAPIView(APIView):
    """API views for exporting financial data in various formats."""

    permission_classes = [permissions.IsAuthenticated, IsSchoolOwnerOrAdmin]


@api_view(["GET", "POST"])
@permission_classes([permissions.IsAuthenticated, IsSchoolOwnerOrAdmin])
def export_transactions_fixed(request: Request) -> Response:
    """
    Export transaction data as CSV.

    GET Query parameters or POST body parameters:
    - format: 'csv' (default)
    - start_date: YYYY-MM-DD (optional, defaults to 30 days ago)
    - end_date: YYYY-MM-DD (optional, defaults to today)
    """
    try:
        # Support both GET query params and POST body data
        params = request.data if request.method == "POST" else request.query_params
        # Handle format parameter
        format_type = params.get("format", "csv")
        if format_type != "csv":
            return Response({"error": "Only CSV format is supported"}, status=status.HTTP_400_BAD_REQUEST)

        # Handle date parameters with validation
        start_date_str = params.get("start_date")
        end_date_str = params.get("end_date")

        # Default date range (30 days)
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=30)

        if start_date_str and end_date_str:
            try:
                requested_start = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                requested_end = datetime.strptime(end_date_str, "%Y-%m-%d").date()

                if requested_start > requested_end:
                    return Response(
                        {"error": "Invalid date range: start date must be before end date"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                start_date = requested_start
                end_date = requested_end
            except ValueError:
                return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=status.HTTP_400_BAD_REQUEST)

        # Get transactions
        transactions = (
            PurchaseTransaction.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            .select_related("student")
            .order_by("-created_at")
        )

        # Create CSV response
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="transactions_{start_date}_to_{end_date}.csv"'

        writer = csv.writer(response)

        # Write header
        writer.writerow(
            [
                "Transaction ID",
                "Student Email",
                "Student Name",
                "Amount",
                "Transaction Type",
                "Payment Status",
                "Created At",
                "Stripe Payment Intent ID",
            ]
        )

        # Write data rows
        for transaction in transactions:
            writer.writerow(
                [
                    transaction.id,
                    transaction.student.email,
                    transaction.student.name,
                    str(transaction.amount),
                    transaction.get_transaction_type_display(),
                    transaction.get_payment_status_display(),
                    transaction.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                    transaction.stripe_payment_intent_id or "",
                ]
            )

        return response

    except Exception as e:
        import traceback

        logger = logging.getLogger(__name__)
        logger.error(f"Transaction export error: {e}")
        logger.error(f"Full traceback: {traceback.format_exc()}")
        return Response(
            {"error": f"Failed to export transactions: {e!s}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated, IsSchoolOwnerOrAdmin])
def test_export_endpoint(request: Request) -> Response:
    """Test endpoint to debug query parameter routing issues."""
    return Response(
        {
            "message": "Test endpoint working",
            "query_params": dict(request.query_params),
            "method": request.method,
            "path": request.path,
        }
    )


class ExportStudentBalancesAPIView(APIView):
    """API view for exporting student balance data as Excel."""

    permission_classes = [permissions.IsAuthenticated, IsSchoolOwnerOrAdmin]

    def _export_balances(self, params):
        """Shared logic for exporting student balances."""
        format_type = params.get("format", "excel")
        if format_type != "excel":
            return Response({"error": "Only Excel format is supported"}, status=status.HTTP_400_BAD_REQUEST)

        if not HAS_OPENPYXL:
            return Response(
                {"error": "Excel export not available - openpyxl not installed"}, status=status.HTTP_501_NOT_IMPLEMENTED
            )

        # Get student balances
        balances = StudentAccountBalance.objects.select_related("student").all()

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Balances"

        # Write header
        headers = [
            "Student Email",
            "Student Name",
            "Hours Purchased",
            "Hours Consumed",
            "Balance Amount",
            "Created At",
            "Last Updated",
        ]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Write data
        for row_num, balance in enumerate(balances, 2):
            ws.cell(row=row_num, column=1, value=balance.student.email)
            ws.cell(row=row_num, column=2, value=balance.student.name)
            ws.cell(row=row_num, column=3, value=str(balance.hours_purchased))
            ws.cell(row=row_num, column=4, value=str(balance.hours_consumed))
            ws.cell(row=row_num, column=5, value=str(balance.balance_amount))
            ws.cell(row=row_num, column=6, value=balance.created_at.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row_num, column=7, value=balance.updated_at.strftime("%Y-%m-%d %H:%M:%S"))

        # Create response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="student_balances.xlsx"'

        wb.save(response)
        return response

    def get(self, request: Request) -> Response:
        """
        Export student balance data as Excel.

        Query parameters:
        - format: 'excel' (required)
        """
        try:
            return self._export_balances(request.query_params)
        except Exception as e:
            logger.error(f"Student balances export error: {e}")
            return Response(
                {"error": "Failed to export student balances"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def post(self, request: Request) -> Response:
        """
        Export student balance data as Excel.

        Body parameters:
        - format: 'excel' (required)
        """
        try:
            return self._export_balances(request.data)
        except Exception as e:
            logger.error(f"Student balances export error: {e}")
            return Response(
                {"error": "Failed to export student balances"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ExportTeacherSessionsAPIView(APIView):
    """API view for exporting teacher session data as HTML report."""

    permission_classes = [permissions.IsAuthenticated, IsSchoolOwnerOrAdmin]

    def _export_teacher_sessions(self, params):
        """Shared logic for exporting teacher sessions."""
        format_type = params.get("format", "html")
        if format_type != "html":
            return Response({"error": "Only HTML format is supported"}, status=status.HTTP_400_BAD_REQUEST)

        teacher_id = params.get("teacher_id")
        month = params.get("month")
        year = params.get("year")

        if not all([teacher_id, month, year]):
            return Response(
                {"error": "teacher_id, month, and year parameters are required"}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            teacher_id = int(teacher_id)
            month = int(month)
            year = int(year)
        except ValueError:
            return Response({"error": "Invalid parameter format"}, status=status.HTTP_400_BAD_REQUEST)

        # Simple HTML response for teacher sessions report
        response = HttpResponse(content_type="text/html")
        response["Content-Disposition"] = (
            f'attachment; filename="teacher_{teacher_id}_sessions_{year}_{month:02d}.html"'
        )

        # Generate HTML report content
        html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Teacher {teacher_id} Sessions Report - {year}/{month:02d}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        .header {{ text-align: center; margin-bottom: 20px; }}
        .report-title {{ font-size: 24px; color: #007bff; }}
        .info {{ margin: 10px 0; }}
    </style>
</head>
<body>
    <div class="header">
        <div class="report-title">Teacher Sessions Report</div>
        <div class="info">Teacher ID: {teacher_id}</div>
        <div class="info">Period: {month:02d}/{year}</div>
    </div>
    <p>This is a sample HTML report for teacher sessions. In a full implementation,
    this would contain detailed session data, statistics, and analytics.</p>
</body>
</html>"""

        response.write(html_content.encode("utf-8"))

        return response

    def get(self, request: Request) -> Response:
        """
        Export teacher session data as HTML report.

        Query parameters:
        - format: 'html' (required)
        - teacher_id: Teacher ID (required)
        - month: Month number (1-12) (required)
        - year: Year (required)
        """
        try:
            return self._export_teacher_sessions(request.query_params)
        except Exception as e:
            logger.error(f"Teacher sessions export error: {e}")
            return Response(
                {"error": "Failed to export teacher sessions"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def post(self, request: Request) -> Response:
        """
        Export teacher session data as HTML report.

        Body parameters:
        - format: 'html' (required)
        - teacher_id: Teacher ID (required)
        - month: Month number (1-12) (required)
        - year: Year (required)
        """
        try:
            return self._export_teacher_sessions(request.data)
        except Exception as e:
            logger.error(f"Teacher sessions export error: {e}")
            return Response(
                {"error": "Failed to export teacher sessions"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ReceiptAPIView(APIView):
    """API views for receipt generation and management."""

    permission_classes = [permissions.IsAuthenticated]


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def generate_receipt(request: Request) -> Response:
    """Generate a receipt for a completed transaction."""
    try:
        serializer = ReceiptGenerationRequestSerializer(data=request.data, context={"request": request})

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        transaction_id = serializer.validated_data["transaction_id"]

        # Get the transaction
        try:
            transaction = PurchaseTransaction.objects.get(id=transaction_id, student=request.user)
        except PurchaseTransaction.DoesNotExist:
            return Response({"error": "Transaction not found"}, status=status.HTTP_404_NOT_FOUND)

        # Check if transaction is completed
        if transaction.payment_status != TransactionPaymentStatus.COMPLETED:
            return Response(
                {"error": "Receipt can only be generated for completed transactions"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Generate receipt using the service
        receipt_data = ReceiptGenerationService.generate_receipt(transaction_id)

        return Response(
            {
                "success": True,
                "receipt_id": receipt_data["receipt_id"],
                "receipt_number": receipt_data["receipt_number"],
                "amount": float(transaction.amount),
            },
            status=status.HTTP_201_CREATED,
        )

    except Exception as e:
        logger.error(f"Receipt generation error: {e}")
        return Response({"error": "Failed to generate receipt"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def list_receipts(request: Request) -> Response:
    """List receipts for the authenticated user."""
    try:
        receipts = Receipt.objects.filter(student=request.user).order_by("-created_at")

        # Add pagination if needed
        page_size = 20
        page = int(request.query_params.get("page", 1))
        start = (page - 1) * page_size
        end = start + page_size

        paginated_receipts = receipts[start:end]

        serializer = ReceiptSerializer(paginated_receipts, many=True)

        # Add download URLs
        for receipt_data, receipt in zip(serializer.data, paginated_receipts, strict=False):
            receipt_data["download_url"] = f"/api/finances/receipts/{receipt.id}/download/"

        return Response({"results": serializer.data, "count": receipts.count(), "page": page, "page_size": page_size})

    except Exception as e:
        logger.error(f"Receipt listing error: {e}")
        return Response({"error": "Failed to list receipts"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def download_receipt(request: Request, pk: int) -> Response:
    """Download a specific receipt HTML."""
    try:
        receipt = Receipt.objects.get(id=pk, student=request.user)

        if not receipt.pdf_file:
            return Response({"error": "HTML file not available"}, status=status.HTTP_404_NOT_FOUND)

        response = HttpResponse(receipt.pdf_file.read(), content_type="text/html")
        response["Content-Disposition"] = f'attachment; filename="{receipt.receipt_number}.html"'

        return response

    except Receipt.DoesNotExist:
        return Response({"error": "Receipt not found"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        logger.error(f"Receipt download error: {e}")
        return Response({"error": "Failed to download receipt"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
